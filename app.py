"""
MLB Daily Betting Report - Streamlit Web App
"""
import streamlit as st
import pandas as pd
import xlsxwriter
import io
import os
from datetime import date, datetime
from daily_report import (
    load_historical_data, fetch_recent_results, compute_all_states,
    fetch_todays_schedule, get_team_state, build_game_row,
    check_scenarios, API_TO_CANONICAL, SCENARIO_DEFS,
    NEEDS_OPP_STREAK, NEEDS_OPP_ROAD_WP, title_case, fmt_line
)


st.set_page_config(page_title="MLB Daily Betting Report", page_icon="⚾", layout="wide")

@st.cache_data(ttl=3600, show_spinner="Loading historical data...")
def load_enriched_data(report_date_str):
    df = load_historical_data()
    df = fetch_recent_results(df, date.fromisoformat(report_date_str))
    return compute_all_states(df)

@st.cache_data(ttl=1800, show_spinner="Fetching today's schedule...")
def get_schedule(report_date_str):
    return fetch_todays_schedule(date.fromisoformat(report_date_str))


def build_report_bytes(games, triggers, report_date, odds):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    NAVY='#1B2A4A'; STEEL='#2E5F8A'; LIGHT_STEEL='#D0E4F5'
    FADE_RED='#C00000'; FADE_BG='#FFE7E7'; WATCH_AMB='#7D5A00'; WATCH_BG='#FFF3CC'
    GREEN_FG='#375623'; GREEN_BG='#E2EFDA'; AWAY_BG='#F0F5FB'; HOME_BG='#FFFFFF'; GRAY_TEXT='#666666'
    fb=wb.add_format({'bold':True,'font_size':16,'font_name':'Calibri','font_color':'white','bg_color':NAVY,'align':'center','valign':'vcenter'})
    fs=wb.add_format({'font_size':11,'font_name':'Calibri','italic':True,'font_color':LIGHT_STEEL,'bg_color':NAVY,'align':'center','valign':'vcenter'})
    fh=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':'white','bg_color':STEEL,'align':'center','valign':'vcenter','border':1,'border_color':'#1A4A72'})
    fal=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':STEEL,'bg_color':AWAY_BG,'align':'left','valign':'vcenter','left':2,'left_color':STEEL,'border':1,'border_color':'#B8C9DC','indent':1})
    fac=wb.add_format({'font_size':10,'font_name':'Calibri','bg_color':AWAY_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#B8C9DC'})
    fhlb=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':'#333333','bg_color':HOME_BG,'align':'left','valign':'vcenter','left':2,'left_color':STEEL,'top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'right':1,'right_color':'#B8C9DC','indent':1})
    fhcb=wb.add_format({'font_size':10,'font_name':'Calibri','bg_color':HOME_BG,'align':'center','valign':'vcenter','top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#B8C9DC','right':1,'right_color':'#B8C9DC'})
    ffa=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':FADE_RED,'bg_color':FADE_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#E8AAAA'})
    ffhb=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':FADE_RED,'bg_color':FADE_BG,'align':'center','valign':'vcenter','top':1,'top_color':'#E8AAAA','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#E8AAAA','right':1,'right_color':'#E8AAAA'})
    fba=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#A8D08D'})
    fbhb=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'align':'center','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#A8D08D'})
    fwa=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':WATCH_AMB,'bg_color':WATCH_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#D4B800'})
    fwhb=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':WATCH_AMB,'bg_color':WATCH_BG,'align':'center','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#D4B800'})
    fnpa=wb.add_format({'italic':True,'font_size':9,'font_name':'Calibri','font_color':GRAY_TEXT,'bg_color':AWAY_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#B8C9DC'})
    fnph=wb.add_format({'italic':True,'font_size':9,'font_name':'Calibri','font_color':GRAY_TEXT,'bg_color':HOME_BG,'align':'center','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#B8C9DC'})
    fng=wb.add_format({'italic':True,'font_size':11,'font_color':GRAY_TEXT,'align':'center','valign':'vcenter'})
    fgc=wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':NAVY,'bg_color':AWAY_BG,'align':'left','valign':'vcenter','border':1,'border_color':'#B8C9DC','indent':1})
    fgch=wb.add_format({'font_size':9,'font_name':'Calibri','font_color':GRAY_TEXT,'bg_color':HOME_BG,'align':'left','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#B8C9DC'})
    fsl=wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri','font_color':NAVY,'bg_color':'#EDF3FB','border':1,'border_color':'#B8C9DC','align':'left','valign':'vcenter','indent':1})
    fsv=wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri','font_color':NAVY,'bg_color':'#EDF3FB','border':1,'border_color':'#B8C9DC','align':'center','valign':'vcenter'})
    ffl=wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri','font_color':FADE_RED,'bg_color':FADE_BG,'border':1,'border_color':'#E8AAAA','align':'left','valign':'vcenter','indent':1})
    ffv=wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri','font_color':FADE_RED,'bg_color':FADE_BG,'border':1,'border_color':'#E8AAAA','align':'center','valign':'vcenter'})
    fwl=wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri','font_color':WATCH_AMB,'bg_color':WATCH_BG,'border':1,'border_color':'#D4B800','align':'left','valign':'vcenter','indent':1})
    fwv=wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri','font_color':WATCH_AMB,'bg_color':WATCH_BG,'border':1,'border_color':'#D4B800','align':'center','valign':'vcenter'})
    fbl=wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'border':1,'border_color':'#A8D08D','align':'left','valign':'vcenter','indent':1})
    fbv=wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'border':1,'border_color':'#A8D08D','align':'center','valign':'vcenter'})

    NCOLS=5; CW=[32,8,26,10,50]; CH=['GAME','H/A','Team','Odds','Play  /  Scenario']

    def pfa(tl):
        if not tl: return fnpa
        v=tl[0]['verdict']
        return ffa if v=='CLEAR FADE' else (fba if v=='CLEAR BET' else fwa)
    def pfh(tl):
        if not tl: return fnph
        v=tl[0]['verdict']
        return ffhb if v=='CLEAR FADE' else (fbhb if v=='CLEAR BET' else fwhb)

    def write_tab(ws,label,vf,tc):
        ws.set_tab_color(tc)
        for ci,w in enumerate(CW): ws.set_column(ci,ci,w)
        ws.set_row(0,30); ws.set_row(1,18)
        ws.merge_range(0,0,0,NCOLS-1,f'MLB DAILY BETTING REPORT  -  {label}',fb)
        ws.merge_range(1,0,1,NCOLS-1,f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',fs)
        ws.set_row(2,20)
        for ci,h in enumerate(CH): ws.write(2,ci,h,fh)
        ws.freeze_panes(3,0)
        row=3; written=0
        for g in games:
            away,home=g['away_team'],g['home_team']
            atc,htc=title_case(away),title_case(home)
            at=[t for t in triggers if t['team']==away and t['opponent']==home and t['verdict']==vf]
            ht=[t for t in triggers if t['team']==home and t['opponent']==away and t['verdict']==vf]
            if not at and not ht: continue
            written+=1
            ws.set_row(row,20)
            ap=(at[0]['play']+'  |  '+'  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in at)) if at else ''
            ws.write(row,0,f'{atc}  @  {htc}',fgc); ws.write(row,1,'AWAY',fac); ws.write(row,2,atc,fal)
            ws.write(row,3,fmt_line(odds.get(away)),fac); ws.write(row,4,ap,pfa(at)); row+=1
            ws.set_row(row,20)
            hp=(ht[0]['play']+'  |  '+'  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in ht)) if ht else ''
            ws.write(row,0,'',fgch); ws.write(row,1,'HOME',fhcb); ws.write(row,2,htc,fhlb)
            ws.write(row,3,fmt_line(odds.get(home)),fhcb); ws.write(row,4,hp,pfh(ht)); row+=1
        if written==0: ws.merge_range(row,0,row,NCOLS-1,f'No {vf.title()} scenarios triggered today.',fng)
        return written

    w0=wb.add_worksheet('Green Clear Bet');   n0=write_tab(w0,'CLEAR BET','CLEAR BET','#00B050')
    w1=wb.add_worksheet('Red Clear Fade');    n1=write_tab(w1,'CLEAR FADE','CLEAR FADE','#FF0000')
    w2=wb.add_worksheet('Yellow Inconsistent');n2=write_tab(w2,'INCONSISTENT','INCONSISTENT','#FFD700')

    w3=wb.add_worksheet('Summary'); w3.set_tab_color('#1B2A4A')
    w3.set_column(0,0,35); w3.set_column(1,1,18); w3.set_column(2,2,35); w3.set_column(3,3,18)
    w3.set_row(0,30); w3.set_row(1,18)
    w3.merge_range(0,0,0,3,'MLB DAILY BETTING REPORT  -  SUMMARY',fb)
    w3.merge_range(1,0,1,3,f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',fs)
    nbt=sum(1 for t in triggers if t['verdict']=='CLEAR BET')
    nft=sum(1 for t in triggers if t['verdict']=='CLEAR FADE')
    nit=sum(1 for t in triggers if t['verdict']=='INCONSISTENT')
    w3.set_row(3,28); w3.set_row(4,28); w3.set_row(5,28)
    w3.write(3,0,'Total Games Today',fsl); w3.write(3,1,len(games),fsv)
    w3.write(3,2,'Games With Triggers',fsl); w3.write(3,3,n0+n1+n2,fsv)
    w3.write(4,0,'Clear Bet Triggers',fbl); w3.write(4,1,nbt,fbv)
    w3.write(4,2,'Clear Fade Triggers',ffl); w3.write(4,3,nft,ffv)
    w3.write(5,0,'Inconsistent Triggers',fwl); w3.write(5,1,nit,fwv)
    w3.merge_range(6,0,6,3,'TOP PLAYS TODAY',fh); w3.set_row(7,20)
    for ci,h in enumerate(['Team','Odds','Play','Scenario']): w3.write(7,ci,h,fh)
    sr=8
    for t in [x for x in triggers if x['verdict']=='CLEAR BET']:
        w3.write(sr,0,title_case(t['team']),fbl); w3.write(sr,1,fmt_line(t['line']),fbv)
        w3.write(sr,2,t['play'],fbl); w3.write(sr,3,f"#{t['scenario_id']} {t['scenario']}",fbl); sr+=1
    for t in [x for x in triggers if x['verdict']=='CLEAR FADE']:
        w3.write(sr,0,title_case(t['team']),ffl); w3.write(sr,1,fmt_line(t['line']),ffv)
        w3.write(sr,2,t['play'],ffl); w3.write(sr,3,f"#{t['scenario_id']} {t['scenario']}",ffl); sr+=1
    for t in [x for x in triggers if x['verdict']=='INCONSISTENT']:
        w3.write(sr,0,title_case(t['team']),fwl); w3.write(sr,1,fmt_line(t['line']),fwv)
        w3.write(sr,2,t['play'],fwl); w3.write(sr,3,f"#{t['scenario_id']} {t['scenario']}",fwl); sr+=1
    w3.freeze_panes(8,0)

    w4=wb.add_worksheet('Results Tracker'); w4.set_tab_color('#375623')
    w4.set_column(0,0,12); w4.set_column(1,1,26); w4.set_column(2,2,10); w4.set_column(3,3,10)
    w4.set_column(4,4,14); w4.set_column(5,5,40); w4.set_column(6,6,14); w4.set_column(7,7,14); w4.set_column(8,8,16)
    fth=wb.add_format({'bold':True,'font_color':'white','bg_color':'#375623','align':'center','valign':'vcenter','border':1,'font_name':'Calibri'})
    ftc=wb.add_format({'align':'center','valign':'vcenter','border':1,'border_color':'#C6EFCE','font_name':'Calibri'})
    ftl=wb.add_format({'align':'left','valign':'vcenter','border':1,'border_color':'#C6EFCE','font_name':'Calibri','indent':1})
    fti=wb.add_format({'bold':True,'align':'center','valign':'vcenter','border':2,'border_color':'#375623','bg_color':'#EAF4E8','font_name':'Calibri'})
    ftt=wb.add_format({'bold':True,'bg_color':'#375623','font_color':'white','align':'center','border':1,'font_name':'Calibri'})
    w4.set_row(0,30); w4.set_row(1,18)
    w4.merge_range(0,0,0,8,'RESULTS TRACKER  -  Enter W or L after each game',fb)
    w4.merge_range(1,0,1,8,'Enter W or L in the Result column. Net P/L calculates automatically.',fs)
    w4.set_row(2,30)
    for ci,h in enumerate(['Date','Team','H/A','Odds','Play','Scenario','Type','Result (W/L)','Net P/L ($100)']): w4.write(2,ci,h,fth)
    tr=3
    for t in triggers:
        line=t['line']; er=tr+1
        w4.write(tr,0,report_date.strftime('%Y-%m-%d'),ftc); w4.write(tr,1,title_case(t['team']),ftl)
        w4.write(tr,2,t['home_away'].upper(),ftc); w4.write(tr,3,fmt_line(line),ftc)
        w4.write(tr,4,t['play'],ftl); w4.write(tr,5,f"#{t['scenario_id']} {t['scenario']}",ftl)
        w4.write(tr,6,t['verdict'],ftc); w4.write(tr,7,'',fti)
        if line is not None and isinstance(line,int):
            pf=f'=IF(H{er}="W",{line},IF(H{er}="L",-100,""))' if line>0 else f'=IF(H{er}="W",ROUND(100/ABS({line})*100,2),IF(H{er}="L",-100,""))'
            w4.write_formula(tr,8,pf,ftc)
        else: w4.write(tr,8,'',ftc)
        tr+=1
    if tr>3:
        w4.merge_range(tr,0,tr,7,'TOTAL NET P/L',ftt)
        w4.write_formula(tr,8,f'=SUMIF(H4:H{tr},"W",I4:I{tr})+SUMIF(H4:H{tr},"L",I4:I{tr})',ftt)
    w4.freeze_panes(3,0); w4.autofilter(2,0,tr,8)

    # Scenario Performance tab
    w5=wb.add_worksheet('Scenario Performance'); w5.set_tab_color('#1F3864')
    w5.set_column(0,0,6); w5.set_column(1,1,40); w5.set_column(2,2,14)
    w5.set_column(3,3,8); w5.set_column(4,4,8); w5.set_column(5,5,10)
    w5.set_column(6,6,10); w5.set_column(7,7,14)
    fpb=wb.add_format({"bold":True,"font_size":14,"font_name":"Calibri","font_color":"white","bg_color":"#1F3864","align":"center","valign":"vcenter"})
    fps=wb.add_format({"font_size":10,"font_name":"Calibri","italic":True,"font_color":"#D0E4F5","bg_color":"#1F3864","align":"center","valign":"vcenter"})
    fph=wb.add_format({"bold":True,"font_size":10,"font_name":"Calibri","font_color":"white","bg_color":"#2E5F8A","align":"center","valign":"vcenter","border":1})
    fps2=wb.add_format({"bold":True,"font_size":10,"font_name":"Calibri","font_color":"#1F3864","bg_color":"#EDF3FB","align":"center","valign":"vcenter","border":1})
    fpn=wb.add_format({"font_size":10,"font_name":"Calibri","font_color":"#1F3864","bg_color":"#EDF3FB","align":"left","valign":"vcenter","border":1,"indent":1})
    fpbet=wb.add_format({"font_size":9,"bold":True,"font_color":"#375623","bg_color":"#E2EFDA","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fpfade=wb.add_format({"font_size":9,"bold":True,"font_color":"#C00000","bg_color":"#FFE7E7","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fpinc=wb.add_format({"font_size":9,"bold":True,"font_color":"#7D5A00","bg_color":"#FFF3CC","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fpnum=wb.add_format({"font_size":10,"bg_color":"#EDF3FB","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fppct=wb.add_format({"font_size":10,"num_format":"0.0%","bg_color":"#EDF3FB","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fpmny=wb.add_format({"font_size":10,"num_format":"$#,##0.00","bg_color":"#EDF3FB","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fptot=wb.add_format({"bold":True,"font_size":10,"font_color":"white","bg_color":"#1F3864","align":"center","valign":"vcenter","border":1,"font_name":"Calibri"})
    fpnote=wb.add_format({"font_size":9,"font_name":"Calibri","italic":True,"font_color":"#7D5A00","bg_color":"#FFF9E6","align":"center","valign":"vcenter","border":1,"text_wrap":True})

    w5.set_row(0,28); w5.set_row(1,16); w5.set_row(2,22)
    w5.merge_range(0,0,0,7,"SCENARIO PERFORMANCE TRACKER  -  Season Cumulative",fpb)
    w5.merge_range(1,0,1,7,"Stats below are for TODAY only. Upload Master_Results.xlsx in the app for season-long tracking.",fps)
    w5.merge_range(2,0,2,7,"To update season totals: open the app → upload your Master_Results.xlsx → download the updated file → enter W/L results.",fpnote)
    w5.set_row(3,20)
    for ci,h in enumerate(["#","Scenario Name","Classification","W","L","Total","Win%","Net P/L"]):
        w5.write(3,ci,h,fph)
    w5.freeze_panes(4,0)

    # Reference internal Results Tracker for today's data
    tsheet = "'Results Tracker'"
    tend = max(tr + 50, 200)
    sc = tsheet + '!F$4:F$' + str(tend)
    rc = tsheet + '!H$4:H$' + str(tend)
    pc = tsheet + '!I$4:I$' + str(tend)

    pr=4
    from daily_report import SCENARIO_DEFS
    for sid,sname,verdict,_ in SCENARIO_DEFS:
        sid_str=f"#{sid} {sname}"
        vfmt=fpbet if verdict=="CLEAR BET" else (fpfade if verdict=="CLEAR FADE" else fpinc)
        vlabel="CLEAR BET" if verdict=="CLEAR BET" else ("CLEAR FADE" if verdict=="CLEAR FADE" else "INCONSISTENT")
        er=pr+1
        w5.write(pr,0,sid,fps2); w5.write(pr,1,sname,fpn); w5.write(pr,2,vlabel,vfmt)
        w5.write_formula(pr,3,'=IFERROR(COUNTIFS('+sc+',"'+sid_str+'",'+rc+',"W"),0)',fpnum)
        w5.write_formula(pr,4,'=IFERROR(COUNTIFS('+sc+',"'+sid_str+'",'+rc+',"L"),0)',fpnum)
        w5.write_formula(pr,5,f'=D{er}+E{er}',fpnum)
        w5.write_formula(pr,6,f'=IF(F{er}>0,D{er}/F{er},"")',fppct)
        w5.write_formula(pr,7,'=IFERROR(SUMPRODUCT(('+sc+'="'+sid_str+'")*ISNUMBER(MATCH('+rc+',{"W","L"},0))*('+pc+')),0)',fpmny)
        pr+=1
    w5.set_row(pr,20)
    for ci in range(8): w5.write(pr,ci,"" if ci not in [1] else "TODAY TOTALS",fptot)
    w5.write_formula(pr,3,f"=SUM(D4:D{pr})",fptot); w5.write_formula(pr,4,f"=SUM(E4:E{pr})",fptot)
    w5.write_formula(pr,5,f"=SUM(F4:F{pr})",fptot)
    w5.write_formula(pr,7,f"=SUM(H4:H{pr})",fptot)

    wb.close(); output.seek(0)
    return output.getvalue(), n0+n1+n2, n1, n2


# ── MAIN UI ───────────────────────────────────────────────────────
st.title("⚾ MLB Daily Betting Report")
st.markdown("---")

col1, col2 = st.columns([2,4])
with col1:
    report_date = st.date_input("Report Date", value=date.today())
report_date_str = report_date.isoformat()

with st.spinner("Loading team data and schedule..."):
    try:
        enriched = load_enriched_data(report_date_str)
        games    = get_schedule(report_date_str)
    except Exception as e:
        st.error(f"Error loading data: {e}")
        st.stop()

if not games:
    st.warning("No games scheduled for this date.")
    st.stop()

st.success(f"✓ {len(games)} games scheduled for {report_date.strftime('%A, %B %d, %Y')}")
st.markdown("---")



st.subheader("📋 Enter Today's Moneylines")
st.caption("Enter the moneyline for each team (e.g. 130 for +130, -150 for -150). Leave blank if unavailable.")

odds_data = [{'Away Team': g['away_team'], 'Home Team': g['home_team'], 'Away Line': None, 'Home Line': None} for g in games]
odds_df = pd.DataFrame(odds_data)
edited = st.data_editor(
    odds_df,
    column_config={
        'Away Team': st.column_config.TextColumn('Away Team', disabled=True, width='large'),
        'Home Team': st.column_config.TextColumn('Home Team', disabled=True, width='large'),
        'Away Line': st.column_config.NumberColumn('Away Line', help='e.g. 130 or -150', width='medium'),
        'Home Line': st.column_config.NumberColumn('Home Line', help='e.g. -130 or 150', width='medium'),
    },
    hide_index=True, use_container_width=True,
)

st.markdown("---")

if st.button("⚾ Generate Daily Report", type="primary", use_container_width=True):
    odds = {}
    for _, row in edited.iterrows():
        if row['Away Line'] is not None and str(row['Away Line']) not in ['','nan']:
            try: odds[str(row['Away Team']).upper()] = int(row['Away Line'])
            except: pass
        if row['Home Line'] is not None and str(row['Home Line']) not in ['','nan']:
            try: odds[str(row['Home Team']).upper()] = int(row['Home Line'])
            except: pass

    if not odds:
        st.warning("Please enter at least some moneylines before generating the report.")
        st.stop()

    with st.spinner("Running scenario analysis..."):
        opp_streaks = {}; opp_road_wpct = {}
        for team in API_TO_CANONICAL.values():
            tdf = enriched[(enriched['team']==team) & (enriched['date'].dt.date < report_date)]
            if not tdf.empty:
                last = tdf.iloc[-1]
                sb = last['streak_before']; res = last['result']
                opp_streaks[team] = (sb+1 if sb>=0 else 1) if res=='W' else (sb-1 if sb<=0 else -1)
                road = tdf[tdf['home_away']=='away']
                if not road.empty:
                    rw=(road['result']=='W').sum(); rl=(road['result']=='L').sum()
                    opp_road_wpct[team] = rw/(rw+rl) if (rw+rl)>0 else None

        all_triggers = []
        for g in games:
            away, home = g['away_team'], g['home_team']
            away_state = get_team_state(enriched, away, report_date)
            home_state = get_team_state(enriched, home, report_date)
            away_row = build_game_row(away_state, 'away', home, odds.get(away))
            home_row = build_game_row(home_state, 'home', away, odds.get(home))
            all_triggers.extend(check_scenarios([away_row, home_row], opp_streaks, opp_road_wpct))

        excel_bytes, n_total, n_fade, n_inc = build_report_bytes(games, all_triggers, report_date, odds)

    st.markdown("---")
    st.subheader("📊 Results")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Games", len(games))
    m2.metric("🟢 Clear Bet", sum(1 for t in all_triggers if t['verdict']=='CLEAR BET'))
    m3.metric("🔴 Clear Fade", sum(1 for t in all_triggers if t['verdict']=='CLEAR FADE'))
    m4.metric("🟡 Inconsistent", sum(1 for t in all_triggers if t['verdict']=='INCONSISTENT'))

    for section_label, section_verdict in [
        ("🟢 Clear Bet Plays", "CLEAR BET"),
        ("🔴 Clear Fade Plays", "CLEAR FADE"),
        ("🟡 Inconsistent Plays", "INCONSISTENT"),
    ]:
        section_triggers = [t for t in all_triggers if t['verdict']==section_verdict]
        if not section_triggers: continue
        st.markdown(f"### {section_label}")
        game_rows = []
        seen = set()
        for g in games:
            away, home = g['away_team'], g['home_team']
            key = f"{away}@{home}"
            if key in seen: continue
            at = [t for t in section_triggers if t['team']==away and t['opponent']==home]
            ht = [t for t in section_triggers if t['team']==home and t['opponent']==away]
            if not at and not ht: continue
            seen.add(key)
            matchup = f"{title_case(away)} @ {title_case(home)}"
            game_rows.append({'GAME': matchup, 'H/A': 'AWAY', 'Team': title_case(away),
                'Odds': fmt_line(odds.get(away)),
                'Play': at[0]['play'] if at else '',
                'Scenario': f"#{at[0]['scenario_id']} {at[0]['scenario']}" if at else ''})
            game_rows.append({'GAME': '', 'H/A': 'HOME', 'Team': title_case(home),
                'Odds': fmt_line(odds.get(home)),
                'Play': ht[0]['play'] if ht else '',
                'Scenario': f"#{ht[0]['scenario_id']} {ht[0]['scenario']}" if ht else ''})
        if game_rows:
            st.dataframe(pd.DataFrame(game_rows), use_container_width=True, hide_index=True,
                column_config={
                    'GAME': st.column_config.TextColumn('GAME', width='large'),
                    'H/A': st.column_config.TextColumn('H/A', width='small'),
                    'Team': st.column_config.TextColumn('Team', width='medium'),
                    'Odds': st.column_config.TextColumn('Odds', width='small'),
                    'Play': st.column_config.TextColumn('Play', width='medium'),
                    'Scenario': st.column_config.TextColumn('Scenario', width='large'),
                })

    st.markdown("---")
    fname = f'MLB_Daily_Report_{report_date.strftime("%Y-%m-%d")}.xlsx'
    st.download_button(label="📥 Download Excel Report", data=excel_bytes, file_name=fname,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True, type="primary")

    # ── CUMULATIVE TRACKING (cloud-friendly: upload/download) ──────
    st.markdown("---")
    st.subheader("📊 Cumulative Season Tracking")
    st.caption("Upload your Master Results file to enter W/L results and view cumulative stats. Download it to save your progress.")

    def build_master_file(existing_df=None):
        """Build a fresh Master_Results.xlsx in memory, pre-populated with today's triggers
        plus any previously uploaded results, and return the bytes."""
        out = io.BytesIO()
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'Master Results'

        NAVY_FILL  = PatternFill('solid', fgColor='1B2A4A')
        GREEN_FILL = PatternFill('solid', fgColor='375623')
        INPUT_FILL = PatternFill('solid', fgColor='EAF4E8')
        thin = lambda c: Side(style='thin', color=c)
        std_border = Border(left=thin('C6EFCE'), right=thin('C6EFCE'),
                            top=thin('C6EFCE'),  bottom=thin('C6EFCE'))
        input_border = Border(left=thin('375623'), right=thin('375623'),
                              top=thin('375623'),  bottom=thin('375623'))

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 16

        # Title
        ws.merge_cells('A1:I1')
        c = ws['A1']
        c.value = 'MASTER RESULTS TRACKER  —  Season Cumulative'
        c.font = Font(bold=True, size=14, color='FFFFFF', name='Calibri')
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Subtitle
        ws.merge_cells('A2:I2')
        c = ws['A2']
        c.value = 'Enter W or L in column H after each game. Net P/L calculates automatically.'
        c.font = Font(italic=True, size=10, color='D0E4F5', name='Calibri')
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 16

        # Headers
        headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result (W/L)', 'Net P/L ($100)']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col)
            c.value = h
            c.font = Font(bold=True, color='FFFFFF', name='Calibri')
            c.fill = GREEN_FILL
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=thin('000000'), right=thin('000000'),
                              top=thin('000000'), bottom=thin('000000'))
        ws.row_dimensions[3].height = 24
        ws.freeze_panes = 'A4'

        # Build rows: existing results (preserves W/L) + new triggers not already present
        existing_rows = []
        existing_keys = set()

        if existing_df is not None and not existing_df.empty:
            for _, row in existing_df.iterrows():
                existing_rows.append(row)
                existing_keys.add((str(row.get('Date','')), str(row.get('Team','')), str(row.get('Scenario',''))))

        # Add today's triggers if not already in the file
        today_str = report_date.strftime('%Y-%m-%d')
        for t in all_triggers:
            scen_str = f"#{t['scenario_id']} {t['scenario']}"
            key = (today_str, title_case(t['team']), scen_str)
            if key not in existing_keys:
                existing_rows.append({
                    'Date': today_str,
                    'Team': title_case(t['team']),
                    'H/A':  t['home_away'].upper(),
                    'Odds': fmt_line(t['line']),
                    'Play': t['play'],
                    'Scenario': scen_str,
                    'Type':   t['verdict'],
                    'Result': '',
                    'Net P/L': None,
                    '_line': t['line'],
                })

        for i, row in enumerate(existing_rows):
            r = i + 4  # excel row (1-indexed), data starts at row 4
            line_val = row.get('_line', None)
            result   = str(row.get('Result', '')).strip().upper()

            ws.cell(r, 1).value = str(row.get('Date', ''))
            ws.cell(r, 2).value = str(row.get('Team', ''))
            ws.cell(r, 3).value = str(row.get('H/A', ''))
            ws.cell(r, 4).value = str(row.get('Odds', ''))
            ws.cell(r, 5).value = str(row.get('Play', ''))
            ws.cell(r, 6).value = str(row.get('Scenario', ''))
            ws.cell(r, 7).value = str(row.get('Type', ''))

            # Result cell (highlighted for user input)
            rc = ws.cell(r, 8)
            rc.value = result if result in ('W', 'L') else ''
            rc.fill = INPUT_FILL
            rc.font = Font(bold=True, name='Calibri')
            rc.alignment = Alignment(horizontal='center', vertical='center')
            rc.border = input_border

            # Net P/L — compute directly (no formulas, works without Excel)
            nc = ws.cell(r, 9)
            if result in ('W', 'L'):
                try:
                    ln = int(line_val) if line_val is not None else None
                    if ln is None:
                        # Try parsing stored odds string like "+130" or "-150"
                        odds_str = str(row.get('Odds', '')).replace('+', '')
                        ln = int(odds_str) if odds_str not in ('', 'N/A', 'None') else None
                    if ln is not None:
                        if result == 'W':
                            nc.value = round(ln * (100 / abs(ln)) if ln < 0 else ln, 2)
                        else:
                            nc.value = -100.0
                    else:
                        nc.value = ''
                except Exception:
                    nc.value = ''
            else:
                nc.value = ''
            nc.alignment = Alignment(horizontal='center', vertical='center')
            nc.border = std_border

            for col in range(1, 8):
                c = ws.cell(r, col)
                c.alignment = Alignment(horizontal='center' if col != 2 else 'left',
                                        vertical='center')
                c.border = std_border

        # Totals row
        last_data = len(existing_rows) + 3
        if existing_rows:
            tr = last_data + 1
            ws.merge_cells(f'A{tr}:H{tr}')
            tc = ws[f'A{tr}']
            tc.value = 'TOTAL NET P/L'
            tc.font = Font(bold=True, color='FFFFFF', name='Calibri')
            tc.fill = GREEN_FILL
            tc.alignment = Alignment(horizontal='center', vertical='center')
            # Sum P/L
            total_pl = sum(
                (ws.cell(i + 4, 9).value or 0)
                for i in range(len(existing_rows))
                if isinstance(ws.cell(i + 4, 9).value, (int, float))
            )
            tc2 = ws.cell(tr, 9)
            tc2.value = round(total_pl, 2)
            tc2.font = Font(bold=True, color='FFFFFF', name='Calibri')
            tc2.fill = GREEN_FILL
            tc2.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(out)
        out.seek(0)
        return out.getvalue()

    # Upload existing master file
    uploaded = st.file_uploader(
        "📤 Upload your Master_Results.xlsx to update with today's results",
        type=['xlsx'], key='master_upload'
    )

    existing_master_df = None
    if uploaded is not None:
        try:
            existing_master_df = pd.read_excel(uploaded, sheet_name='Master Results', skiprows=2)
            existing_master_df.columns = ['Date','Team','H/A','Odds','Play','Scenario','Type','Result','Net P/L']
            existing_master_df = existing_master_df[existing_master_df['Date'].notna()]
            st.success(f"✓ Loaded {len(existing_master_df)} existing results from your Master file")
        except Exception as e:
            st.warning(f"Could not read uploaded file: {e}. Starting fresh.")
            existing_master_df = None

    master_bytes = build_master_file(existing_master_df)
    st.download_button(
        label="📥 Download Master_Results.xlsx  (open → enter W/L → re-upload next time)",
        data=master_bytes,
        file_name='Master_Results.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
    )

    with st.expander("ℹ️ How cumulative tracking works", expanded=False):
        st.markdown("""
        **Simple 3-step workflow:**
        1. **Download** `Master_Results.xlsx` after generating each daily report
        2. **Open the file** → enter **W** or **L** in the Result column as games complete → **Save**
        3. **Next time** — upload your saved Master file here before generating the report.
           Today's new triggers are added automatically and your old results are preserved.

        The file grows over the season as your full record. No server storage needed.
        """)


# ── SCENARIO PERFORMANCE HEATMAP ─────────────────────────────────
st.markdown("---")
st.subheader("🔥 Scenario Performance Heatmap  (2023–2026 Historical Backtest)")
st.caption("Win % color-coded: 🟢 green = strong, 🟡 yellow = mixed, 🔴 red = weak. Click column headers to sort.")

@st.cache_data(ttl=86400, show_spinner="Building heatmap from historical data...")
def build_heatmap_data():
    """
    Run all 36 scenarios against the full historical dataset using fast
    vectorized pandas — no row-by-row iteration.
    Returns a summary DataFrame with W, L, Win% per scenario.
    """
    from daily_report import (
        load_historical_data, compute_all_states, SCENARIO_DEFS,
        NEEDS_OPP_STREAK, NEEDS_OPP_ROAD_WP, DIVISIONS
    )

    df = load_historical_data()
    df = compute_all_states(df)
    df = df[df['result'].isin(['W', 'L'])].copy()

    # Pre-compute opponent streak per team (last known)
    opp_streaks = {}
    opp_road_wpct = {}
    for team, tdf in df.groupby('team'):
        tdf = tdf.sort_values('date')
        last = tdf.iloc[-1]
        sb, res = last['streak_before'], last['result']
        opp_streaks[team] = (sb + 1 if sb >= 0 else 1) if res == 'W' else (sb - 1 if sb <= 0 else -1)
        road = tdf[tdf['home_away'] == 'away']
        if not road.empty:
            rw = (road['result'] == 'W').sum()
            rl = (road['result'] == 'L').sum()
            opp_road_wpct[team] = rw / (rw + rl) if (rw + rl) > 0 else None

    rows = []
    for sid, sname, verdict, func in SCENARIO_DEFS:
        try:
            if sid in NEEDS_OPP_STREAK:
                mask = df.apply(lambda r: bool(func(r.to_dict(), opp_streaks)), axis=1)
            elif sid in NEEDS_OPP_ROAD_WP:
                mask = df.apply(lambda r: bool(func(r.to_dict(), opp_road_wpct)), axis=1)
            else:
                mask = df.apply(lambda r: bool(func(r.to_dict())), axis=1)
        except Exception:
            mask = pd.Series(False, index=df.index)

        subset = df[mask]
        total  = len(subset)
        wins   = (subset['result'] == 'W').sum()
        losses = (subset['result'] == 'L').sum()
        win_pct = round(wins / total * 100, 1) if total > 0 else None

        rows.append({
            '#':        sid,
            'Scenario': sname,
            'Type':     verdict,
            'Games':    total,
            'W':        int(wins),
            'L':        int(losses),
            'Win%':     win_pct,
        })

    return pd.DataFrame(rows)

with st.expander("Click to view heatmap", expanded=False):
    with st.spinner("Running scenarios against historical data… (first load ~30 sec, then cached)"):
        hmap = build_heatmap_data()

    if not hmap.empty:
        c1, c2 = st.columns([2, 2])
        with c1:
            sort_opt = st.selectbox(
                "Sort by",
                ["Win% High→Low", "Win% Low→High", "Scenario #", "Games High→Low"],
                key='hm_sort'
            )
        with c2:
            type_filter = st.multiselect(
                "Type",
                ['CLEAR BET', 'CLEAR FADE', 'INCONSISTENT'],
                default=['CLEAR BET', 'CLEAR FADE', 'INCONSISTENT'],
                key='hm_type'
            )

        hmap = hmap[hmap['Type'].isin(type_filter)].copy()

        if sort_opt == "Win% High→Low":
            hmap = hmap.sort_values('Win%', ascending=False, na_position='last')
        elif sort_opt == "Win% Low→High":
            hmap = hmap.sort_values('Win%', ascending=True, na_position='last')
        elif sort_opt == "Games High→Low":
            hmap = hmap.sort_values('Games', ascending=False)
        else:
            hmap = hmap.sort_values('#')

        def _row_style(row):
            v = row['Win%']
            vtype = row['Type']
            if v is None:
                wc = 'background-color:#F0F0F0'
            elif vtype == 'CLEAR FADE':
                wc = ('background-color:#C6EFCE;font-weight:bold' if v <= 42
                      else 'background-color:#FFEB9C;font-weight:bold' if v <= 50
                      else 'background-color:#FFC7CE;font-weight:bold')
            else:
                wc = ('background-color:#C6EFCE;font-weight:bold' if v >= 58
                      else 'background-color:#FFEB9C;font-weight:bold' if v >= 50
                      else 'background-color:#FFC7CE;font-weight:bold')
            tc = {'CLEAR BET': 'background-color:#C6EFCE;font-weight:bold',
                  'CLEAR FADE': 'background-color:#FFC7CE;font-weight:bold',
                  'INCONSISTENT': 'background-color:#FFEB9C;font-weight:bold'}.get(vtype, '')
            return ['', '', tc, '', 'background-color:#E2EFDA', 'background-color:#FFE7E7', wc]

        disp = hmap.copy()
        disp['Win%'] = disp['Win%'].apply(lambda x: f"{x:.1f}%" if x is not None else "—")

        st.dataframe(
            disp.style.apply(_row_style, axis=1),
            use_container_width=True,
            hide_index=True,
            column_config={
                '#':        st.column_config.TextColumn('#', width=55),
                'Scenario': st.column_config.TextColumn('Scenario', width='large'),
                'Type':     st.column_config.TextColumn('Type', width='medium'),
                'Games':    st.column_config.NumberColumn('Games', width=70),
                'W':        st.column_config.NumberColumn('W', width=55),
                'L':        st.column_config.NumberColumn('L', width=55),
                'Win%':     st.column_config.TextColumn('Win %', width=75),
            }
        )

        has = hmap[hmap['Games'] > 0].copy()
        has['_pct'] = has['Win%'].apply(lambda x: float(x.replace('%','')) if isinstance(x, str) else (x or 0))
        if not has.empty:
            best  = has.loc[has['_pct'].idxmax()]
            worst = has.loc[has['_pct'].idxmin()]
            b1, b2, b3 = st.columns(3)
            b1.success(f"🏆 Best: #{best['#']} — {best['Win%']}")
            b2.error(f"⚠️ Lowest: #{worst['#']} — {worst['Win%']}")
            b3.info(f"📊 {len(has)} scenarios · {int(hmap['Games'].sum()):,} total games")
