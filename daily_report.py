"""
MLB Daily Betting Report Generator
====================================
Workflow:
  1. Auto-fetches today's MLB schedule from MLB Stats API (free, no key needed)
  2. Auto-updates team game logs with recent results
  3. Client pastes today's moneylines into: daily_odds_input.xlsx
  4. Script checks each game against all 36 scenarios
  5. Outputs: MLB_Daily_Report_YYYY-MM-DD.xlsx  (Clear Fade + Inconsistent tabs)

Run manually each morning, or schedule via Windows Task Scheduler / cron.
"""

import requests
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime, timedelta, date
import re
import os
import sys

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

DATA_FILE      = 'MLB Data 2023-2026.xlsx'   # historical data (kept updated)
ODDS_INPUT     = 'daily_odds_input.xlsx'      # client fills this each morning
OUTPUT_DIR     = '.'                          # where to save daily reports
REPORT_DATE    = date.today()                 # override: date(2026, 7, 7)

# ─────────────────────────────────────────────
# TEAM NAME MAPPING  (MLB API name -> our canonical name)
# ─────────────────────────────────────────────

API_TO_CANONICAL = {
    'Chicago White Sox':    'CHICAGO WHITE SOX',
    'Cleveland Guardians':  'CLEVELAND GUARDIANS',
    'Detroit Tigers':       'DETROIT TIGERS',
    'Kansas City Royals':   'KANSAS CITY ROYALS',
    'Minnesota Twins':      'MINNESOTA TWINS',
    'Baltimore Orioles':    'BALTIMORE ORIOLES',
    'Boston Red Sox':       'BOSTON RED SOX',
    'New York Yankees':     'NEW YORK YANKEES',
    'Tampa Bay Rays':       'TAMPA BAY RAYS',
    'Toronto Blue Jays':    'TORONTO BLUE JAYS',
    'Oakland Athletics':    'ATHLETICS',
    'Athletics':            'ATHLETICS',
    'Houston Astros':       'HOUSTON ASTROS',
    'Los Angeles Angels':   'LOS ANGELES ANGELS',
    'Seattle Mariners':     'SEATTLE MARINERS',
    'Texas Rangers':        'TEXAS RANGERS',
    'Chicago Cubs':         'CHICAGO CUBS',
    'Cincinnati Reds':      'CINCINNATI REDS',
    'Milwaukee Brewers':    'MILWAUKEE BREWERS',
    'Pittsburgh Pirates':   'PITTSBURGH PIRATES',
    'St. Louis Cardinals':  'ST. LOUIS CARDINALS',
    'Atlanta Braves':       'ATLANTA BRAVES',
    'Miami Marlins':        'MIAMI MARLINS',
    'New York Mets':        'NEW YORK METS',
    'Philadelphia Phillies':'PHILADELPHIA PHILLIES',
    'Washington Nationals': 'WASHINGTON NATIONALS',
    'Arizona Diamondbacks': 'ARIZONA DIAMONDBACKS',
    'Colorado Rockies':     'COLORADO ROCKIES',
    'Los Angeles Dodgers':  'LOS ANGELES DODGERS',
    'San Diego Padres':     'SAN DIEGO PADRES',
    'San Francisco Giants': 'SAN FRANCISCO GIANTS',
}

TEAMS = list(set(API_TO_CANONICAL.values()))

AL_EAST    = {'BALTIMORE ORIOLES','BOSTON RED SOX','NEW YORK YANKEES','TAMPA BAY RAYS','TORONTO BLUE JAYS'}
AL_CENTRAL = {'CHICAGO WHITE SOX','CLEVELAND GUARDIANS','DETROIT TIGERS','KANSAS CITY ROYALS','MINNESOTA TWINS'}
AL_WEST    = {'ATHLETICS','HOUSTON ASTROS','LOS ANGELES ANGELS','SEATTLE MARINERS','TEXAS RANGERS'}
NL_EAST    = {'ATLANTA BRAVES','MIAMI MARLINS','NEW YORK METS','PHILADELPHIA PHILLIES','WASHINGTON NATIONALS'}
NL_CENTRAL = {'CHICAGO CUBS','CINCINNATI REDS','MILWAUKEE BREWERS','PITTSBURGH PIRATES','ST. LOUIS CARDINALS'}
NL_WEST    = {'ARIZONA DIAMONDBACKS','COLORADO ROCKIES','LOS ANGELES DODGERS','SAN DIEGO PADRES','SAN FRANCISCO GIANTS'}

DIVISIONS = {}
for t in AL_EAST:    DIVISIONS[t] = 'AL_EAST'
for t in AL_CENTRAL: DIVISIONS[t] = 'AL_CENTRAL'
for t in AL_WEST:    DIVISIONS[t] = 'AL_WEST'
for t in NL_EAST:    DIVISIONS[t] = 'NL_EAST'
for t in NL_CENTRAL: DIVISIONS[t] = 'NL_CENTRAL'
for t in NL_WEST:    DIVISIONS[t] = 'NL_WEST'

OPP_NORM = {
    'houston': 'HOUSTON ASTROS', 'san francisco': 'SAN FRANCISCO GIANTS',
    'pittsburgh': 'PITTSBURGH PIRATES', 'minnesota': 'MINNESOTA TWINS',
    'baltimore': 'BALTIMORE ORIOLES', 'philadelphia': 'PHILADELPHIA PHILLIES',
    'boston': 'BOSTON RED SOX', 'ny yankees': 'NEW YORK YANKEES',
    'new york yankees': 'NEW YORK YANKEES', 'tampa bay': 'TAMPA BAY RAYS',
    'toronto': 'TORONTO BLUE JAYS', 'la angels': 'LOS ANGELES ANGELS',
    'los angeles angels': 'LOS ANGELES ANGELS', 'chicago white sox': 'CHICAGO WHITE SOX',
    'chi white sox': 'CHICAGO WHITE SOX', 'seattle': 'SEATTLE MARINERS',
    'oakland': 'ATHLETICS', 'athletics': 'ATHLETICS',
    'cleveland': 'CLEVELAND GUARDIANS', 'detroit': 'DETROIT TIGERS',
    'kansas city': 'KANSAS CITY ROYALS', 'texas': 'TEXAS RANGERS',
    'chicago cubs': 'CHICAGO CUBS', 'chi cubs': 'CHICAGO CUBS',
    'cincinnati': 'CINCINNATI REDS', 'milwaukee': 'MILWAUKEE BREWERS',
    'st. louis': 'ST. LOUIS CARDINALS', 'atlanta': 'ATLANTA BRAVES',
    'miami': 'MIAMI MARLINS', 'ny mets': 'NEW YORK METS',
    'new york mets': 'NEW YORK METS', 'washington': 'WASHINGTON NATIONALS',
    'arizona': 'ARIZONA DIAMONDBACKS', 'colorado': 'COLORADO ROCKIES',
    'la dodgers': 'LOS ANGELES DODGERS', 'los angeles dodgers': 'LOS ANGELES DODGERS',
    'san diego': 'SAN DIEGO PADRES', 'chicago-al': 'CHICAGO WHITE SOX',
    'chicago-nl': 'CHICAGO CUBS', 'new york-al': 'NEW YORK YANKEES',
    'new york-nl': 'NEW YORK METS',
}

def normalize_opponent(raw_opp):
    s = str(raw_opp).replace('\xa0', ' ').strip()
    s = re.sub(r'^(vs|@)\s*', '', s, flags=re.IGNORECASE).strip().lower()
    if s in OPP_NORM:
        return OPP_NORM[s]
    for key, val in OPP_NORM.items():
        if key in s or s in key:
            return val
    return s.upper()

def parse_score(score_str, result):
    try:
        parts = str(score_str).split('-')
        a, b = int(parts[0]), int(parts[1])
        return (a, b) if result == 'W' else (b, a)
    except:
        return None, None

# ─────────────────────────────────────────────
# STEP 1: LOAD HISTORICAL DATA
# ─────────────────────────────────────────────

def load_historical_data():
    wb = load_workbook(DATA_FILE, read_only=True)
    records = []
    TEAM_LIST = [
        'CHICAGO WHITE SOX','CLEVELAND GUARDIANS','DETROIT TIGERS','KANSAS CITY ROYALS','MINNESOTA TWINS',
        'BALTIMORE ORIOLES','BOSTON RED SOX','NEW YORK YANKEES','TAMPA BAY RAYS','TORONTO BLUE JAYS',
        'ATHLETICS','HOUSTON ASTROS','LOS ANGELES ANGELS','SEATTLE MARINERS','TEXAS RANGERS',
        'CHICAGO CUBS','CINCINNATI REDS','MILWAUKEE BREWERS','PITTSBURGH PIRATES','ST. LOUIS CARDINALS',
        'ATLANTA BRAVES','MIAMI MARLINS','NEW YORK METS','PHILADELPHIA PHILLIES','WASHINGTON NATIONALS',
        'ARIZONA DIAMONDBACKS','COLORADO ROCKIES','LOS ANGELES DODGERS','SAN DIEGO PADRES','SAN FRANCISCO GIANTS'
    ]
    for year_str in ['2023','2024','2025','2026']:
        ws = wb[year_str]
        all_rows = list(ws.iter_rows(values_only=True))
        year = int(year_str)
        for t_idx, team in enumerate(TEAM_LIST):
            col = t_idx * 9
            for row in all_rows[2:]:
                date_val = row[col]
                opp_val  = row[col+1]
                result   = row[col+3]
                score    = row[col+4]
                line     = row[col+5]
                ou       = row[col+6]
                total    = row[col+7]
                if date_val is None or result is None:
                    continue
                opp_str = str(opp_val).replace('\xa0',' ')
                home_away = 'away' if opp_str.startswith('@') else 'home'
                opponent = normalize_opponent(opp_str)
                rs, ra = parse_score(score, result)
                records.append({
                    'year': year, 'team': team,
                    'date': pd.Timestamp(date_val),
                    'opponent': opponent, 'home_away': home_away,
                    'result': result, 'score': score,
                    'runs_scored': rs, 'runs_allowed': ra,
                    'line': line, 'ou': ou, 'total': total,
                    'division': DIVISIONS.get(team,'UNKNOWN'),
                })
    df = pd.DataFrame(records)
    df['date'] = pd.to_datetime(df['date'])
    return df.sort_values(['team','date']).reset_index(drop=True)

# ─────────────────────────────────────────────
# STEP 2: FETCH TODAY'S SCHEDULE FROM MLB API
# ─────────────────────────────────────────────

def fetch_todays_schedule(report_date):
    date_str = report_date.strftime('%Y-%m-%d')
    url = f'https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={date_str}'
    try:
        r = requests.get(url, timeout=10)
        data = r.json()
    except Exception as e:
        print(f'ERROR fetching schedule: {e}')
        return []

    games = []
    if not data.get('dates'):
        print(f'No games found for {date_str}')
        return []

    for g in data['dates'][0]['games']:
        away_api = g['teams']['away']['team']['name']
        home_api = g['teams']['home']['team']['name']
        away = API_TO_CANONICAL.get(away_api, away_api.upper())
        home = API_TO_CANONICAL.get(home_api, home_api.upper())
        game_time = g.get('gameDate','')
        games.append({
            'away_team': away,
            'home_team': home,
            'game_time': game_time,
            'game_pk':   g['gamePk'],
        })
    print(f'Schedule: {len(games)} games on {date_str}')
    return games

# ─────────────────────────────────────────────
# STEP 3: FETCH RECENT RESULTS & UPDATE DATA
# ─────────────────────────────────────────────

def fetch_recent_results(df, report_date):
    """
    Fetch game results from MLB Stats API for dates after the last date in df.
    Appends new completed games to df.
    """
    last_date = df['date'].max().date()
    check_date = last_date + timedelta(days=1)
    new_records = []

    while check_date < report_date:
        date_str = check_date.strftime('%Y-%m-%d')
        url = f'https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={date_str}&hydrate=linescore'
        try:
            r = requests.get(url, timeout=10)
            data = r.json()
        except:
            check_date += timedelta(days=1)
            continue

        if not data.get('dates'):
            check_date += timedelta(days=1)
            continue

        for g in data['dates'][0]['games']:
            if g['status']['detailedState'] != 'Final':
                continue
            away_api   = g['teams']['away']['team']['name']
            home_api   = g['teams']['home']['team']['name']
            away_team  = API_TO_CANONICAL.get(away_api, away_api.upper())
            home_team  = API_TO_CANONICAL.get(home_api, home_api.upper())
            away_score = g['teams']['away'].get('score', 0)
            home_score = g['teams']['home'].get('score', 0)
            away_win   = g['teams']['away'].get('isWinner', False)

            year = check_date.year
            dt   = pd.Timestamp(check_date)

            # Away team record
            if away_win:
                rs_a, ra_a, res_a = away_score, home_score, 'W'
                rs_h, ra_h, res_h = home_score, away_score, 'L'
            else:
                rs_a, ra_a, res_a = away_score, home_score, 'L'
                rs_h, ra_h, res_h = home_score, away_score, 'W'

            score_str = f'{max(away_score,home_score)}-{min(away_score,home_score)}'

            for team, opp, ha, rs, ra, res in [
                (away_team, home_team, 'away', rs_a, ra_a, res_a),
                (home_team, away_team, 'home', rs_h, ra_h, res_h),
            ]:
                new_records.append({
                    'year': year, 'team': team,
                    'date': dt, 'opponent': opp,
                    'home_away': ha, 'result': res,
                    'score': score_str,
                    'runs_scored': rs, 'runs_allowed': ra,
                    'line': None, 'ou': None, 'total': None,
                    'division': DIVISIONS.get(team, 'UNKNOWN'),
                })

        check_date += timedelta(days=1)

    if new_records:
        new_df = pd.DataFrame(new_records)
        df = pd.concat([df, new_df], ignore_index=True)
        df = df.sort_values(['team','date']).reset_index(drop=True)
        print(f'Added {len(new_records)} new game records from API')
    else:
        print('Data already up to date')

    return df

# ─────────────────────────────────────────────
# STEP 4: COMPUTE TEAM STATES
# ─────────────────────────────────────────────

def compute_all_states(df):
    TEAM_LIST = sorted(df['team'].unique())
    parts = []
    for team in TEAM_LIST:
        tdf = df[df['team'] == team].copy().sort_values('date').reset_index(drop=True)
        parts.append(compute_team_states(tdf))
    enriched = pd.concat(parts).reset_index(drop=True)
    return enriched

def compute_team_states(rows):
    n = len(rows)
    streak = [0]*n
    prev_opponent = [None]*n
    prev_result = [None]*n
    prev_line = [None]*n
    prev_runs_scored = [None]*n
    prev_runs_allowed = [None]*n
    series_game_num = [1]*n
    series_id = [0]*n
    homestand_game_num = [0]*n
    roadtrip_game_num = [0]*n
    homestand_series_num = [0]*n
    wins_so_far = [0]*n
    games_so_far = [0]*n
    prev2_rs = [None]*n
    prev3_rs = [None]*n
    prev4_rs = [None]*n
    prev2_ra = [None]*n
    prev3_ra = [None]*n
    prev4_ra = [None]*n

    cur_streak = 0; cur_wins = 0; cur_games = 0; s_id = 0
    runs_scored_hist = []; runs_allowed_hist = []
    cur_home_game = 0; cur_road_game = 0; cur_home_series = 0
    last_location = None; cur_series_start = 0
    cur_year = rows.at[0,'year'] if n > 0 else None

    for i in range(n):
        this_year = rows.at[i,'year']
        if this_year != cur_year:
            cur_streak=0; cur_wins=0; cur_games=0
            runs_scored_hist=[]; runs_allowed_hist=[]
            cur_home_game=0; cur_road_game=0; cur_home_series=0
            last_location=None; cur_year=this_year
            prev_result[i]=None; prev_opponent[i]=None
            prev_line[i]=None; prev_runs_scored[i]=None; prev_runs_allowed[i]=None

        wins_so_far[i] = cur_wins
        games_so_far[i] = cur_games

        if i > 0 and prev_result[i] is None:
            prev_result[i]      = rows.at[i-1,'result']
            prev_opponent[i]    = rows.at[i-1,'opponent']
            prev_line[i]        = rows.at[i-1,'line']
            prev_runs_scored[i] = rows.at[i-1,'runs_scored']
            prev_runs_allowed[i]= rows.at[i-1,'runs_allowed']

        streak[i] = cur_streak

        opp = rows.at[i,'opponent']
        if i == 0 or opp != rows.at[i-1,'opponent']:
            s_id += 1; cur_series_start = i
        series_id[i] = s_id
        series_game_num[i] = i - cur_series_start + 1

        loc = rows.at[i,'home_away']
        if loc != last_location:
            cur_home_game=0; cur_road_game=0; cur_home_series=0; last_location=loc
        if loc == 'home':
            cur_home_game += 1
            if i == 0 or rows.at[i-1,'home_away'] != 'home' or opp != rows.at[i-1,'opponent']:
                cur_home_series = cur_home_series+1 if (i>0 and rows.at[i-1,'home_away']=='home') else 1
            homestand_game_num[i] = cur_home_game
            homestand_series_num[i] = cur_home_series
            roadtrip_game_num[i] = 0
        else:
            cur_road_game += 1
            roadtrip_game_num[i] = cur_road_game
            homestand_game_num[i] = 0
            homestand_series_num[i] = 0

        if len(runs_scored_hist) >= 2:
            prev2_rs[i] = list(runs_scored_hist[-2:])
            prev2_ra[i] = list(runs_allowed_hist[-2:])
        if len(runs_scored_hist) >= 3:
            prev3_rs[i] = list(runs_scored_hist[-3:])
            prev3_ra[i] = list(runs_allowed_hist[-3:])
        if len(runs_scored_hist) >= 4:
            prev4_rs[i] = list(runs_scored_hist[-4:])
            prev4_ra[i] = list(runs_allowed_hist[-4:])

        r = rows.at[i,'result']
        runs_scored_hist.append(rows.at[i,'runs_scored'])
        runs_allowed_hist.append(rows.at[i,'runs_allowed'])
        cur_games += 1
        if r == 'W':
            cur_wins += 1
            cur_streak = cur_streak+1 if cur_streak >= 0 else 1
        else:
            cur_streak = cur_streak-1 if cur_streak <= 0 else -1

    sbp = [None] + streak[:-1]

    rows['streak_before']       = streak
    rows['streak_before_prev']  = sbp
    rows['prev_opponent']       = prev_opponent
    rows['prev_result']         = prev_result
    rows['prev_line']           = prev_line
    rows['prev_runs_scored']    = prev_runs_scored
    rows['prev_runs_allowed']   = prev_runs_allowed
    rows['series_game_num']     = series_game_num
    rows['series_id']           = series_id
    rows['homestand_game_num']  = homestand_game_num
    rows['homestand_series_num']= homestand_series_num
    rows['roadtrip_game_num']   = roadtrip_game_num
    rows['wins_before']         = wins_so_far
    rows['games_before']        = games_so_far
    rows['prev2_runs_scored']   = prev2_rs
    rows['prev3_runs_scored']   = prev3_rs
    rows['prev4_runs_scored']   = prev4_rs
    rows['prev2_runs_allowed']  = prev2_ra
    rows['prev3_runs_allowed']  = prev3_ra
    rows['prev4_runs_allowed']  = prev4_ra

    rows['winpct_before'] = rows.apply(
        lambda r: r['wins_before']/r['games_before'] if r['games_before']>0 else None, axis=1)

    results_list = rows['result'].tolist()
    year_list    = rows['year'].tolist()
    l10 = []
    for i in range(len(results_list)):
        window = []
        for j in range(i-1, max(-1,i-11), -1):
            if year_list[j] != year_list[i]: break
            window.append(results_list[j])
        l10.append(window.count('W'))
    rows['last10_wins'] = l10

    series_totals = rows.groupby('series_id')['series_id'].transform('count')
    rows['series_total'] = series_totals.values

    rows['rt_segment'] = (rows['home_away'] != rows['home_away'].shift()).cumsum()
    rt_totals = rows[rows['home_away']=='away'].groupby('rt_segment')['rt_segment'].transform('count')
    rows['roadtrip_total'] = 0
    rows.loc[rows['home_away']=='away','roadtrip_total'] = rt_totals.values

    hs_totals = rows[rows['home_away']=='home'].groupby('rt_segment')['rt_segment'].transform('count')
    rows['homestand_total'] = 0
    rows.loc[rows['home_away']=='home','homestand_total'] = hs_totals.values

    return rows


# ─────────────────────────────────────────────
# STEP 5: GET CURRENT STATE FOR TODAY'S GAMES
# ─────────────────────────────────────────────

def get_team_state(enriched, team, report_date):
    """Get the most recent state for a team as of report_date (before today's game)."""
    tdf = enriched[
        (enriched['team'] == team) &
        (enriched['date'].dt.date < report_date)
    ].sort_values('date')

    if tdf.empty:
        return None

    last = tdf.iloc[-1]
    # Compute current streak after last game
    sb = last['streak_before']
    res = last['result']
    if res == 'W':
        cur_streak = sb + 1 if sb >= 0 else 1
    else:
        cur_streak = sb - 1 if sb <= 0 else -1

    return {
        'team':               team,
        'streak_before':      cur_streak,
        'streak_before_prev': last['streak_before'],
        'prev_result':        last['result'],
        'prev_opponent':      last['opponent'],
        'prev_line':          last['line'],
        'prev_runs_scored':   last['runs_scored'],
        'prev_runs_allowed':  last['runs_allowed'],
        'series_game_num':    last['series_game_num'],
        'series_total':       last['series_total'],
        'series_id':          last['series_id'],
        'homestand_game_num': last['homestand_game_num'],
        'homestand_series_num': last['homestand_series_num'],
        'roadtrip_game_num':  last['roadtrip_game_num'],
        'roadtrip_total':     last['roadtrip_total'],
        'homestand_total':    last['homestand_total'],
        'wins_before':        last['wins_before'] + (1 if last['result']=='W' else 0),
        'games_before':       last['games_before'] + 1,
        'last10_wins':        last['last10_wins'],
        'prev2_runs_scored':  last['prev2_runs_scored'],
        'prev3_runs_scored':  last['prev3_runs_scored'],
        'prev4_runs_scored':  last['prev4_runs_scored'],
        'prev2_runs_allowed': last['prev2_runs_allowed'],
        'prev3_runs_allowed': last['prev3_runs_allowed'],
        'prev4_runs_allowed': last['prev4_runs_allowed'],
        'winpct_before':      (last['wins_before']+(1 if last['result']=='W' else 0)) / (last['games_before']+1),
        'division':           DIVISIONS.get(team,'UNKNOWN'),
    }

def build_game_row(state, home_away, opponent, line):
    """Build a pseudo-row for scenario filtering from team state + today's game info."""
    if state is None:
        return None

    # Determine series position for today's game
    # If same opponent as last game, we're continuing the series
    last_opp = state['prev_opponent']
    if last_opp == opponent:
        series_game = state['series_game_num'] + 1
        series_total = state['series_total']  # approximate
    else:
        series_game = 1
        series_total = 3  # assume 3-game series default

    # Homestand/road trip position
    if home_away == 'home':
        hg = state['homestand_game_num'] + 1 if state['homestand_game_num'] > 0 else 1
        rg = 0
        hs_series = state['homestand_series_num']
        if last_opp != opponent and state['homestand_game_num'] > 0:
            hs_series += 1
        elif state['homestand_game_num'] == 0:
            hs_series = 1
        rt_total = 0
    else:
        rg = state['roadtrip_game_num'] + 1 if state['roadtrip_game_num'] > 0 else 1
        hg = 0
        hs_series = 0
        rt_total = state['roadtrip_total']

    return {
        'team':               state['team'],
        'opponent':           opponent,
        'home_away':          home_away,
        'line':               line,
        'streak_before':      state['streak_before'],
        'streak_before_prev': state['streak_before_prev'],
        'prev_result':        state['prev_result'],
        'prev_opponent':      state['prev_opponent'],
        'prev_line':          state['prev_line'],
        'prev_runs_scored':   state['prev_runs_scored'],
        'prev_runs_allowed':  state['prev_runs_allowed'],
        'series_game_num':    series_game,
        'series_total':       series_total,
        'homestand_game_num': hg,
        'homestand_series_num': hs_series,
        'roadtrip_game_num':  rg,
        'roadtrip_total':     rt_total,
        'last10_wins':        state['last10_wins'],
        'wins_before':        state['wins_before'],
        'games_before':       state['games_before'],
        'winpct_before':      state['winpct_before'],
        'prev2_runs_scored':  state['prev2_runs_scored'],
        'prev3_runs_scored':  state['prev3_runs_scored'],
        'prev4_runs_scored':  state['prev4_runs_scored'],
        'prev2_runs_allowed': state['prev2_runs_allowed'],
        'prev3_runs_allowed': state['prev3_runs_allowed'],
        'prev4_runs_allowed': state['prev4_runs_allowed'],
        'division':           state['division'],
    }


# ─────────────────────────────────────────────
# STEP 6: SCENARIO FILTERS (all 36)
# ─────────────────────────────────────────────

def same_opp(r): return r['prev_opponent'] == r['opponent']
def diff_opp(r): return r['prev_opponent'] != r['opponent'] and r['prev_opponent'] is not None
def is_home(r):  return r['home_away'] == 'home'
def is_away(r):  return r['home_away'] == 'away'
def prev_won(r): return r['prev_result'] == 'W'
def prev_lost(r):return r['prev_result'] == 'L'

def run_diff_prev(r):
    rs, ra = r['prev_runs_scored'], r['prev_runs_allowed']
    return (rs-ra) if (rs is not None and ra is not None) else None

def all_rs_le(lst, t):
    return lst is not None and all(v is not None and v<=t for v in lst)

def all_ra_ge(lst, t):
    return lst is not None and all(v is not None and v>=t for v in lst)

def last_game_series(r): return r['series_game_num'] == r['series_total']
def first_game_series(r): return r['series_game_num'] == 1
def second_game_series(r): return r['series_game_num'] == 2
def last_game_roadtrip(r): return is_away(r) and r['roadtrip_game_num']==r['roadtrip_total'] and r['roadtrip_total']>0
def first_game_homestand(r): return is_home(r) and r['homestand_game_num']==1
def div_match(t1,t2): return DIVISIONS.get(t1)==DIVISIONS.get(t2)

def s01(r):
    if not prev_won(r) or not same_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 8

def s02(r):
    if not prev_won(r) or not diff_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 7

def s03(r):
    if not prev_won(r) or not diff_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 11

def s04(r):
    pl, cl = r['prev_line'], r['line']
    if pl is None or cl is None: return False
    return prev_won(r) and same_opp(r) and pl >= 150 and cl <= 125

def s05(r):
    pl, cl = r['prev_line'], r['line']
    if pl is None or cl is None: return False
    return prev_lost(r) and diff_opp(r) and pl <= -150 and cl >= 100

def s06(r):
    if not is_away(r) or not prev_lost(r): return False
    sbp = r['streak_before_prev']
    return sbp is not None and sbp >= 5

def s07(r):
    cl = r['line']
    if cl is None: return False
    sbp = r['streak_before_prev']
    return prev_won(r) and diff_opp(r) and sbp is not None and sbp <= -6 and cl <= 175

def s08(r):
    cl = r['line']
    if cl is None: return False
    sbp = r['streak_before_prev']
    return is_home(r) and prev_won(r) and same_opp(r) and sbp is not None and sbp <= -4 and cl >= -200

def s09(r, opp_streaks):
    if not is_away(r) or r['streak_before'] >= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s10(r, opp_streaks):
    cl = r['line']
    if cl is None or not is_home(r) or r['streak_before'] >= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and -3 <= os <= -1 and cl <= -150

def s11(r, opp_streaks):
    if not is_home(r) or r['streak_before'] > -4: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and -3 <= os <= -1

def s12(r, opp_streaks):
    if not is_away(r) or r['streak_before'] <= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s13(r, opp_streaks):
    if not is_home(r) or r['streak_before'] not in [1,2]: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s14(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -2: return False
    return all_rs_le(r['prev2_runs_scored'], 2) and cl >= 150

def s15(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -4: return False
    return all_rs_le(r['prev4_runs_scored'], 3) and cl >= -105

def s16(r):
    cl = r['line']
    if cl is None or not is_home(r) or r['streak_before'] > -4: return False
    return all_rs_le(r['prev4_runs_scored'], 2) and cl >= -200

def s17(r):
    if not is_away(r) or r['streak_before'] > -2: return False
    return all_ra_ge(r['prev2_runs_allowed'], 9)

def s18(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -3: return False
    return all_ra_ge(r['prev3_runs_allowed'], 7) and cl <= 200

def s19(r):
    if not is_away(r) or r['streak_before'] > -4: return False
    return all_ra_ge(r['prev4_runs_allowed'], 6)

def s20(r): return r['streak_before'] >= 3 and diff_opp(r)
def s21(r): return r['streak_before'] <= -3 and last_game_series(r)

def s22(r):
    cl = r['line']
    return cl is not None and is_away(r) and first_game_series(r) and -130 <= cl <= -111

def s23(r):
    cl = r['line']
    return cl is not None and is_away(r) and last_game_series(r) and 101 <= cl <= 187

def s24(r):
    cl = r['line']
    return cl is not None and last_game_roadtrip(r) and 101 <= cl <= 187

def s25(r):
    cl = r['line']
    return (cl is not None and is_home(r) and first_game_series(r)
            and r['homestand_series_num'] >= 2 and -109 <= cl <= 120)

def s26(r):
    cl = r['line']
    return cl is not None and first_game_homestand(r) and -180 <= cl <= -111

def s27(r):
    cl = r['line']
    return (cl is not None and is_home(r) and last_game_series(r)
            and div_match(r['team'], r['opponent']) and cl <= -180)

def s28(r):
    cl = r['line']
    prs = r['prev_runs_scored']
    return (cl is not None and is_away(r) and first_game_series(r)
            and prev_won(r) and prs is not None and prs >= 6 and cl <= 174)

def s29(r):
    prs = r['prev_runs_scored']
    return (is_away(r) and last_game_series(r) and prev_lost(r)
            and prs is not None and prs <= 4)

def s30(r):
    cl = r['line']
    wp = r['winpct_before']
    return (cl is not None and first_game_homestand(r)
            and wp is not None and wp < 0.500 and 110 <= cl <= 170)

def s31(r):
    cl = r['line']
    prs, pra = r['prev_runs_scored'], r['prev_runs_allowed']
    if cl is None or not prev_lost(r) or not same_opp(r): return False
    if prs is None or pra is None: return False
    return (pra - prs) == 1 and 100 <= cl <= 150

def s32(r, opp_road_wpct):
    cl = r['line']
    if cl is None or not is_home(r) or not prev_lost(r): return False
    if r['last10_wins'] > 5 or cl < 100: return False
    if not div_match(r['team'], r['opponent']): return False
    owp = opp_road_wpct.get(r['opponent'])
    return owp is not None and owp <= 0.400

def s33(r):
    cl, pl = r['line'], r['prev_line']
    if cl is None or pl is None: return False
    if not is_home(r) or not second_game_series(r): return False
    wp = r['winpct_before']
    return (cl >= 100 and wp is not None and wp < 0.500
            and r['prev_result'] == 'L' and pl < 0)

def s34(r):
    cl = r['line']
    if cl is None or not is_home(r): return False
    gb, wb = r['games_before'], r['wins_before']
    if gb == 0 or wb*2 != gb: return False
    return r['streak_before'] >= 2 and cl >= -190

def s35(r):
    cl, pl = r['line'], r['prev_line']
    if cl is None or pl is None: return False
    return second_game_series(r) and cl < 0 and r['prev_result']=='W' and pl < 0

def s36(r):
    if r['series_total'] != 3 or not last_game_series(r): return False
    return r['prev_result'] == 'L' and r['streak_before'] == -1


SCENARIO_DEFS = [
    ('01','BLOWOUT #1 - MJ',                           'CLEAR BET',   s01),
    ('02','BLOWOUT #2 - MJ',                           'CLEAR FADE',  s02),
    ('03','BLOWOUT #3 - MJ',                           'INCONSISTENT',s03),
    ('04','THE BIG UPSET #1 - MJ',                     'INCONSISTENT',s04),
    ('05','THE BIG UPSET #2 - MJ',                     'CLEAR FADE',  s05),
    ('06','THE SNAPPED WINNING STREAK - MJ',           'CLEAR FADE',  s06),
    ('07','THE SNAPPED LOSING STREAK #1 - MJ',         'INCONSISTENT',s07),
    ('08','THE SNAPPED LOSING STREAK #2 - MJ',         'INCONSISTENT',s08),
    ('09','THE COLD TEAM VS HOT TEAM - MJ',            'INCONSISTENT',s09),
    ('10','THE COLD TEAMS MATCHUP #1 - MJ',            'INCONSISTENT',s10),
    ('11','THE COLD TEAMS MATCHUP #2 - MJ',            'INCONSISTENT',s11),
    ('12','THE HOT TEAMS MATCHUP #1 - MJ',             'CLEAR FADE',  s12),
    ('13','THE HOT TEAMS MATCHUP #2 - MJ',             'INCONSISTENT',s13),
    ('14','THE SCORING DROUGHT #1 - MJ',               'CLEAR FADE',  s14),
    ('15','THE SCORING DROUGHT #2 - MJ',               'CLEAR FADE',  s15),
    ('16','THE SCORING DROUGHT #3 - MJ',               'CLEAR FADE',  s16),
    ('17','THE PUMMELED PITCHERS #1 - MJ',             'INCONSISTENT',s17),
    ('18','THE PUMMELED PITCHERS #2 - MJ',             'CLEAR FADE',  s18),
    ('19','THE PUMMELED PITCHERS #3 - MJ',             'CLEAR FADE',  s19),
    ('20','3-GAME WIN STREAK / NEW SERIES - SM',       'CLEAR BET',   s20),
    ('21','3-GAME LOSING STREAK / LAST SERIES - SM',   'CLEAR BET',   s21),
    ('22','SMALL ROAD FAVORITE / NEW SERIES - SM',     'INCONSISTENT',s22),
    ('23','ROAD DOG / LAST GAME OF SERIES - SM',       'INCONSISTENT',s23),
    ('24','ROAD DOG / LAST GAME OF ROAD TRIP - SM',    'CLEAR BET',   s24),
    ('25','HOME DOG / MID-HOMESTAND NEW OPP - SM',     'CLEAR FADE',  s25),
    ('26','SMALL HOME FAV / FIRST HOMESTAND - SM',     'CLEAR FADE',  s26),
    ('27','BIG HOME FAV / LAST DIV SERIES - SM',       'INCONSISTENT',s27),
    ('28','HOT ROAD TEAM / FIRST GAME SERIES - SM',    'INCONSISTENT',s28),
    ('29','COLD ROAD TEAM / LAST GAME SERIES - SM',    'INCONSISTENT',s29),
    ('30','HOME DOG OFF LONG ROAD TRIP',               'CLEAR FADE',  s30),
    ('31','DOG OFF CLOSE LOSS',                        'INCONSISTENT',s31),
    ('32','COLD HOME DOG / DIV. GAME',                 'INCONSISTENT',s32),
    ('33','HOME DOG / GAME 2 OF SERIES',               'INCONSISTENT',s33),
    ('34','HOME FAVORITE AT EXACTLY .500',             'INCONSISTENT',s34),
    ('35','SERIES FACTORS #1',                         'CLEAR FADE',  s35),
    ('36','SERIES FACTOR #2',                          'CLEAR FADE',  s36),
]

# Scenarios needing opponent streak (pass extra arg)
NEEDS_OPP_STREAK = {'09','10','11','12','13'}
NEEDS_OPP_ROAD_WP = {'32'}


# ─────────────────────────────────────────────
# STEP 7: ODDS INPUT FILE HANDLER
# ─────────────────────────────────────────────

def create_odds_template(games, report_date):
    """
    Create a simple Excel template for client to fill in today's odds.
    Columns: Away Team | Home Team | Away Line | Home Line
    """
    fname = ODDS_INPUT
    wb = xlsxwriter.Workbook(fname)
    ws = wb.add_worksheet('Odds Input')

    hdr = wb.add_format({'bold':True,'bg_color':'#2E75B6','font_color':'white',
                          'align':'center','border':1})
    cell = wb.add_format({'align':'center','border':1})
    title = wb.add_format({'bold':True,'font_size':13,'align':'center'})

    ws.merge_range(0,0,0,3, f'DAILY ODDS INPUT — {report_date.strftime("%B %d, %Y")}', title)
    ws.write(1,0,'Away Team',hdr)
    ws.write(1,1,'Home Team',hdr)
    ws.write(1,2,'Away Line (e.g. +130 or -150)',hdr)
    ws.write(1,3,'Home Line (e.g. -130 or +150)',hdr)
    ws.set_column(0,1,28)
    ws.set_column(2,3,28)

    for i, g in enumerate(games):
        ws.write(2+i, 0, g['away_team'], cell)
        ws.write(2+i, 1, g['home_team'], cell)
        ws.write(2+i, 2, '', cell)
        ws.write(2+i, 3, '', cell)

    wb.close()
    print(f'Odds template created: {fname}')
    print(f'Please fill in the moneylines and re-run the script.')

def load_odds(games):
    """Load odds from the filled-in template. Returns dict: team -> line."""
    if not os.path.exists(ODDS_INPUT):
        return {}

    wb = load_workbook(ODDS_INPUT, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    odds = {}
    for row in rows[2:]:
        if row[0] is None: continue
        away_team = str(row[0]).strip().upper()
        home_team = str(row[1]).strip().upper() if row[1] else ''
        away_line = row[2]
        home_line = row[3]
        if away_line is not None:
            try: odds[away_team] = int(away_line)
            except: pass
        if home_line is not None:
            try: odds[home_team] = int(home_line)
            except: pass
    return odds


# ─────────────────────────────────────────────
# STEP 8: RUN SCENARIOS FOR TODAY'S GAMES
# ─────────────────────────────────────────────

def check_scenarios(game_rows, opp_streaks, opp_road_wpct):
    """
    For each team row (away/home), check all 36 scenarios.
    Returns list of triggered scenario results.
    """
    triggers = []
    for row in game_rows:
        if row is None: continue
        team    = row['team']
        opp     = row['opponent']
        ha      = row['home_away']
        line    = row['line']

        for sid, sname, verdict, func in SCENARIO_DEFS:
            try:
                if sid in NEEDS_OPP_STREAK:
                    fired = func(row, opp_streaks)
                elif sid in NEEDS_OPP_ROAD_WP:
                    fired = func(row, opp_road_wpct)
                else:
                    fired = func(row)
            except:
                fired = False

            if fired:
                play = f'FADE {team.title()}' if verdict == 'CLEAR FADE' else f'WATCH {team.title()}'
                triggers.append({
                    'team':     team,
                    'opponent': opp,
                    'home_away': ha,
                    'line':     line,
                    'scenario_id': sid,
                    'scenario': sname,
                    'verdict':  verdict,
                    'play':     play,
                })
    return triggers


# ─────────────────────────────────────────────
# STEP 9: BUILD EXCEL REPORT
# ─────────────────────────────────────────────

def title_case(name):
    """Convert ALL CAPS team name to Title Case."""
    exceptions = {'Sox','Red','White','Blue','Los','San','New','St.'}
    return ' '.join(w.capitalize() for w in name.split())

def fmt_line(line):
    """Format moneyline for display."""
    if line is None or line == 'N/A': return 'N/A'
    return f'+{line}' if isinstance(line, int) and line > 0 else str(line)


def build_report(games, triggers, report_date, odds):
    fname = os.path.join(OUTPUT_DIR, f'MLB_Daily_Report_{report_date.strftime("%Y-%m-%d")}.xlsx')
    wb = xlsxwriter.Workbook(fname)

    # ── COLOUR PALETTE ───────────────────────────────────────────
    NAVY        = '#1B2A4A'
    STEEL       = '#2E5F8A'
    LIGHT_STEEL = '#D0E4F5'
    FADE_RED    = '#C00000'
    FADE_BG     = '#FFE7E7'
    WATCH_AMB   = '#7D5A00'
    WATCH_BG    = '#FFF3CC'
    AWAY_BG     = '#F0F5FB'
    HOME_BG     = '#FFFFFF'
    GAME_HDR_BG = '#1B2A4A'
    SCEN_BG     = '#E8F0FA'
    GRAY_TEXT   = '#666666'
    GREEN_BG    = '#E2EFDA'
    GREEN_FG    = '#375623'

    # ── FORMATS ──────────────────────────────────────────────────
    # Banner
    f_banner = wb.add_format({
        'bold': True, 'font_size': 16, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': NAVY,
        'align': 'center', 'valign': 'vcenter',
    })
    f_subtitle = wb.add_format({
        'font_size': 11, 'font_name': 'Calibri', 'italic': True,
        'font_color': LIGHT_STEEL, 'bg_color': NAVY,
        'align': 'center', 'valign': 'vcenter',
    })
    f_col_hdr = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': STEEL,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#1A4A72',
        'text_wrap': False,
    })
    # Game header bar (spans full row, shows "Game N: Away @ Home")
    f_game_hdr = wb.add_format({
        'bold': True, 'font_size': 11, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': GAME_HDR_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'top': 1, 'top_color': '#0A1828',
        'bottom': 1, 'bottom_color': '#0A1828',
    })
    # Away/Home team label cells
    f_away_label = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': STEEL, 'bg_color': AWAY_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'border': 1, 'border_color': '#B8C9DC',
        'indent': 1,
    })
    f_home_label = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': '#333333', 'bg_color': HOME_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'border': 1, 'border_color': '#B8C9DC',
        'indent': 1,
    })
    f_away_cell = wb.add_format({
        'font_size': 10, 'font_name': 'Calibri',
        'bg_color': AWAY_BG, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    f_home_cell = wb.add_format({
        'font_size': 10, 'font_name': 'Calibri',
        'bg_color': HOME_BG, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    # FADE play — red
    f_fade_away = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': FADE_RED, 'bg_color': FADE_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#E8AAAA',
    })
    f_fade_home = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': FADE_RED, 'bg_color': FADE_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#E8AAAA',
    })
    # WATCH play — amber
    f_watch = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': WATCH_AMB, 'bg_color': WATCH_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D4B800',
    })
    # No play — grey dash
    f_no_play = wb.add_format({
        'italic': True, 'font_size': 9, 'font_name': 'Calibri',
        'font_color': GRAY_TEXT, 'bg_color': AWAY_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    f_no_play_h = wb.add_format({
        'italic': True, 'font_size': 9, 'font_name': 'Calibri',
        'font_color': GRAY_TEXT, 'bg_color': HOME_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    # Scenario name cell (light blue bg)
    f_scen_away = wb.add_format({
        'font_size': 9, 'font_name': 'Calibri', 'italic': True,
        'font_color': STEEL, 'bg_color': SCEN_BG,
        'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
        'text_wrap': True, 'indent': 1,
    })
    f_scen_home = wb.add_format({
        'font_size': 9, 'font_name': 'Calibri', 'italic': True,
        'font_color': '#555555', 'bg_color': HOME_BG,
        'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
        'text_wrap': True, 'indent': 1,
    })
    f_no_scen_away = wb.add_format({
        'italic': True, 'font_size': 9, 'font_color': GRAY_TEXT,
        'bg_color': SCEN_BG, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC', 'indent': 1,
    })
    f_no_scen_home = wb.add_format({
        'italic': True, 'font_size': 9, 'font_color': GRAY_TEXT,
        'bg_color': HOME_BG, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC', 'indent': 1,
    })
    f_no_games = wb.add_format({
        'italic': True, 'font_size': 11, 'font_color': GRAY_TEXT,
        'align': 'center', 'valign': 'vcenter',
    })
    # Summary formats
    f_sum_label  = wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':NAVY,'bg_color':'#EDF3FB',
                                   'border':1,'border_color':'#B8C9DC',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_val    = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':NAVY,'bg_color':'#EDF3FB',
                                   'border':1,'border_color':'#B8C9DC',
                                   'align':'center','valign':'vcenter'})
    f_sum_fade_l = wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':FADE_RED,'bg_color':FADE_BG,
                                   'border':1,'border_color':'#E8AAAA',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_fade_v = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':FADE_RED,'bg_color':FADE_BG,
                                   'border':1,'border_color':'#E8AAAA',
                                   'align':'center','valign':'vcenter'})
    f_sum_watch_l= wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':WATCH_AMB,'bg_color':WATCH_BG,
                                   'border':1,'border_color':'#D4B800',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_watch_v= wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':WATCH_AMB,'bg_color':WATCH_BG,
                                   'border':1,'border_color':'#D4B800',
                                   'align':'center','valign':'vcenter'})

    # ── COLUMN LAYOUT ─────────────────────────────────────────────
    # Col 0: #  (game number)        width 5
    # Col 1: H/A badge               width 8
    # Col 2: Team                     width 26
    # Col 3: Odds                     width 10
    # Col 4: Play                     width 22
    # Col 5: Scenario(s)              width 48
    NCOLS = 6
    COL_WIDTHS = [5, 8, 26, 10, 22, 48]
    COL_HEADERS = ['#', 'H/A', 'Team', 'Odds', 'Play', 'Scenario(s)']

    def write_tab(ws, tab_label, verdict_filter, tab_color):
        ws.set_tab_color(tab_color)
        for ci, w in enumerate(COL_WIDTHS):
            ws.set_column(ci, ci, w)

        # ── Banner rows ──
        ws.set_row(0, 30)
        ws.set_row(1, 18)
        ws.merge_range(0, 0, 0, NCOLS-1,
            f'⚾  MLB DAILY BETTING REPORT  —  {tab_label}', f_banner)
        ws.merge_range(1, 0, 1, NCOLS-1,
            f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',
            f_subtitle)

        # ── Column headers ──
        ws.set_row(2, 20)
        for ci, h in enumerate(COL_HEADERS):
            ws.write(2, ci, h, f_col_hdr)

        ws.freeze_panes(3, 0)

        row = 3
        games_written = 0

        for g_idx, g in enumerate(games, 1):
            away = g['away_team']
            home = g['home_team']
            away_tc = title_case(away)
            home_tc = title_case(home)
            away_line = odds.get(away)
            home_line = odds.get(home)

            at = [t for t in triggers if t['team']==away and t['opponent']==home and t['verdict']==verdict_filter]
            ht = [t for t in triggers if t['team']==home and t['opponent']==away and t['verdict']==verdict_filter]
            if not at and not ht:
                continue

            games_written += 1

            # ── Game header bar ──
            ws.set_row(row, 18)
            ws.merge_range(row, 0, row, NCOLS-1,
                f'  Game {games_written}  ▸  {away_tc}  @  {home_tc}', f_game_hdr)
            row += 1

            # ── Away row ──
            ws.set_row(row, 20)
            a_scen = '  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in at) if at else '—'
            a_play = at[0]['play'] if at else '—'
            a_pf   = f_fade_away if (at and at[0]['verdict']=='CLEAR FADE') else (f_watch if at else f_no_play)
            ws.write(row, 0, g_idx,         f_away_cell)
            ws.write(row, 1, 'AWAY',        f_away_cell)
            ws.write(row, 2, away_tc,       f_away_label)
            ws.write(row, 3, fmt_line(away_line), f_away_cell)
            ws.write(row, 4, a_play,        a_pf)
            ws.write(row, 5, a_scen,        f_scen_away if at else f_no_scen_away)
            row += 1

            # ── Home row ──
            ws.set_row(row, 20)
            h_scen = '  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in ht) if ht else '—'
            h_play = ht[0]['play'] if ht else '—'
            h_pf   = f_fade_home if (ht and ht[0]['verdict']=='CLEAR FADE') else (f_watch if ht else f_no_play_h)
            ws.write(row, 0, g_idx,         f_home_cell)
            ws.write(row, 1, 'HOME',        f_home_cell)
            ws.write(row, 2, home_tc,       f_home_label)
            ws.write(row, 3, fmt_line(home_line), f_home_cell)
            ws.write(row, 4, h_play,        h_pf)
            ws.write(row, 5, h_scen,        f_scen_home if ht else f_no_scen_home)
            row += 2  # spacer row between games

        if games_written == 0:
            ws.set_row(row, 30)
            ws.merge_range(row, 0, row, NCOLS-1,
                f'No {verdict_filter.title()} scenarios triggered today.', f_no_games)

        return games_written

    # ── WRITE TABS ────────────────────────────────────────────────
    ws_fade = wb.add_worksheet('🔴 Clear Fade')
    n_fade  = write_tab(ws_fade, 'CLEAR FADE', 'CLEAR FADE', '#FF0000')

    ws_inc  = wb.add_worksheet('🟡 Inconsistent')
    n_inc   = write_tab(ws_inc,  'INCONSISTENT', 'INCONSISTENT', '#FFD700')

    # ── SUMMARY TAB ───────────────────────────────────────────────
    ws_sum = wb.add_worksheet('📊 Summary')
    ws_sum.set_tab_color('#1B2A4A')
    ws_sum.set_column(0, 0, 35)
    ws_sum.set_column(1, 1, 18)
    ws_sum.set_column(2, 2, 35)
    ws_sum.set_column(3, 3, 18)

    ws_sum.set_row(0, 30)
    ws_sum.set_row(1, 18)
    ws_sum.merge_range(0, 0, 0, 3, '⚾  MLB DAILY BETTING REPORT  —  SUMMARY', f_banner)
    ws_sum.merge_range(1, 0, 1, 3,
        f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',
        f_subtitle)

    n_fade_t = sum(1 for t in triggers if t['verdict']=='CLEAR FADE')
    n_inc_t  = sum(1 for t in triggers if t['verdict']=='INCONSISTENT')

    ws_sum.set_row(3, 28)
    ws_sum.set_row(4, 28)
    ws_sum.set_row(5, 28)

    ws_sum.write(3, 0, 'Total Games Today',     f_sum_label)
    ws_sum.write(3, 1, len(games),              f_sum_val)
    ws_sum.write(3, 2, 'Games With Triggers',   f_sum_label)
    ws_sum.write(3, 3, n_fade + n_inc,          f_sum_val)

    ws_sum.write(4, 0, '🔴  Clear Fade Triggers',  f_sum_fade_l)
    ws_sum.write(4, 1, n_fade_t,                   f_sum_fade_v)
    ws_sum.write(4, 2, '🟡  Inconsistent Triggers', f_sum_watch_l)
    ws_sum.write(4, 3, n_inc_t,                    f_sum_watch_v)

    # Top plays section
    ws_sum.set_row(7, 20)
    ws_sum.merge_range(6, 0, 6, 3, 'TOP PLAYS TODAY', f_col_hdr)
    ws_sum.write(7, 0, 'Team',       f_col_hdr)
    ws_sum.write(7, 1, 'Odds',       f_col_hdr)
    ws_sum.write(7, 2, 'Play',       f_col_hdr)
    ws_sum.write(7, 3, 'Scenario',   f_col_hdr)

    sum_row = 8
    for t in [x for x in triggers if x['verdict']=='CLEAR FADE']:
        pf = f_sum_fade_l
        ws_sum.write(sum_row, 0, title_case(t['team']),     pf)
        ws_sum.write(sum_row, 1, fmt_line(t['line']),       f_sum_fade_v)
        ws_sum.write(sum_row, 2, t['play'],                 f_sum_fade_l)
        ws_sum.write(sum_row, 3, f"#{t['scenario_id']} {t['scenario']}", pf)
        sum_row += 1
    for t in [x for x in triggers if x['verdict']=='INCONSISTENT']:
        pf = f_sum_watch_l
        ws_sum.write(sum_row, 0, title_case(t['team']),     pf)
        ws_sum.write(sum_row, 1, fmt_line(t['line']),       f_sum_watch_v)
        ws_sum.write(sum_row, 2, t['play'],                 f_sum_watch_l)
        ws_sum.write(sum_row, 3, f"#{t['scenario_id']} {t['scenario']}", pf)
        sum_row += 1

    if sum_row == 8:
        ws_sum.merge_range(8, 0, 8, 3, 'No scenarios triggered today.', f_no_games)

    ws_sum.freeze_panes(8, 0)

    wb.close()
    print(f'\n✓ Report saved: {fname}')
    print(f'  Clear Fade games: {n_fade} ({n_fade_t} triggers)')
    print(f'  Inconsistent games: {n_inc} ({n_inc_t} triggers)')
    return fname


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == '__main__':
    print(f'=== MLB Daily Report — {REPORT_DATE} ===\n')

    # 1. Load historical data
    print('Loading historical data...')
    df = load_historical_data()

    # 2. Fetch recent results to bring data current
    print('Checking for recent game results...')
    df = fetch_recent_results(df, REPORT_DATE)

    # 3. Compute team states
    print('Computing team states...')
    enriched = compute_all_states(df)

    # 4. Fetch today's schedule
    print('Fetching today\'s schedule...')
    games = fetch_todays_schedule(REPORT_DATE)
    if not games:
        print('No games today. Exiting.')
        sys.exit(0)

    # 5. Load odds — create template if missing
    odds = load_odds(games)
    if not odds:
        print('\nNo odds found. Creating odds input template...')
        create_odds_template(games, REPORT_DATE)
        print('\nFill in the moneylines in daily_odds_input.xlsx then re-run.')
        sys.exit(0)

    # 6. Build opponent streak lookup for today
    opp_streaks = {}
    opp_road_wpct = {}
    for team in API_TO_CANONICAL.values():
        tdf = enriched[(enriched['team']==team) & (enriched['date'].dt.date < REPORT_DATE)]
        if not tdf.empty:
            last = tdf.iloc[-1]
            sb = last['streak_before']
            res = last['result']
            opp_streaks[team] = (sb+1 if sb>=0 else 1) if res=='W' else (sb-1 if sb<=0 else -1)
            # Road win%
            road = tdf[tdf['home_away']=='away']
            if not road.empty:
                rw = (road['result']=='W').sum()
                rl = (road['result']=='L').sum()
                opp_road_wpct[team] = rw/(rw+rl) if (rw+rl)>0 else None

    # 7. Build game rows and check scenarios
    print('Checking scenarios...')
    all_triggers = []
    for g in games:
        away, home = g['away_team'], g['home_team']
        away_line = odds.get(away)
        home_line = odds.get(home)

        away_state = get_team_state(enriched, away, REPORT_DATE)
        home_state = get_team_state(enriched, home, REPORT_DATE)

        away_row = build_game_row(away_state, 'away', home, away_line)
        home_row = build_game_row(home_state, 'home', away, home_line)

        triggers = check_scenarios([away_row, home_row], opp_streaks, opp_road_wpct)
        all_triggers.extend(triggers)

    # 8. Build report
    build_report(games, all_triggers, REPORT_DATE, odds)
